import tkinter as tk
import json
from openpyxl import *
import os
import time


class Edit(tk.Frame):
    def __init__(self, *args):
        tk.Frame.__init__(self, *args)
    def show(self):
        self.lift()

        
class LoanAmt(Edit):
    def __init__(self, *args):
        
        Edit.__init__(self, *args)
        
        
        
        self.labeljName = tk.Label(self, text="Edit Loan Amount",height=4)
        self.labeljName.grid(row=0,column=5,padx=(200,0))

        #--------------------------Veh No.-----------------------------------
        self.labelVehNo = tk.Label(self, text="Veh No:-",height=4)
        self.labelVehNo.grid(row=1,column=3,padx=(200,0))
        self.e11 = tk.Entry(self,width=30)
        self.e11.grid(row=1,columnspan=2,column=4)
        self.tke11 = tk.StringVar()
        self.tke11.set("Nothing is done yet!")

        """self.SetAmount = tk.Label(self, text="Amount:-",height=4)
        self.SetAmount.grid(row=2,column=3,padx=(200,0))
        self.e12 = tk.Entry(self,width=30)
        self.e12.grid(row=2,columnspan=2,column=4,padx=(50,0))"""

        self.buttonFind = tk.Button(self, text="Search", fg="green",width=20)
        self.buttonFind.grid(row=3,column=4,padx=(0,0))
        self.buttonFind.config(command = self.searchLoan)


        

        
    def searchLoan(self):
        self.tke11.set(self.e11.get())
        self.setk11 = self.tke11.get()

        self.wb_cust = load_workbook('G:\\Finance software\\Database\\newLoan.xlsx')

        self.sheet_cust = self.wb_cust.active

        maxcol = self.sheet_cust.max_column
        maxrow = self.sheet_cust.max_row
        c=0
        for i in range(1,maxrow+1):
            cell_obj = self.sheet_cust.cell(row = i,column = 14)
            print(str(cell_obj.value))
            if(str(cell_obj.value)==str(self.setk11)):
                self.row = i
                
                self.SetAmount = tk.Label(self, text="Amount:-",height=4)
                self.SetAmount.grid(row=2,column=3,padx=(200,0))
                self.e12 = tk.Entry(self,width=30)
                self.e12.grid(row=2,columnspan=2,column=4)
                self.tke12 = tk.StringVar()
                self.tke12.set("Nothing is done yet!")

                self.buttonFind = tk.Button(self, text="Submit", fg="green",width=20)
                self.buttonFind.grid(row=3,column=4,padx=(0,0))
                self.buttonFind.config(command = self.changeAmount)


                
                
                
                c=1
                break
        if(c==0):
            self.SetAmount = tk.Label(self, text="Sorry Couldnt Find the Vehicle! Check the Vehicle Number",height=4)
            self.SetAmount.grid(row=2,column=3,padx=(200,0))

    def changeAmount(self):
        #modifying data for the particular loan in the newloan file
        self.wb_cust = load_workbook('G:\\Finance software\\Database\\newLoan.xlsx')

        self.sheet_cust = self.wb_cust.active
        self.tke12.set(self.e12.get())
        self.setk12 = self.tke12.get()
        self.old_amount = self.sheet_cust.cell(row=self.row,column=10).value
        self.sheet_cust.cell(row=self.row,column=10).value = str(self.setk12)
        self.updated_amount = self.sheet_cust.cell(row=self.row,column=10).value


        self.unique_id = self.sheet_cust.cell(row=self.row,column=21).value
        
        self.broker_name = self.sheet_cust.cell(row=self.row,column=11).value

        self.interest_rate = self.sheet_cust.cell(row=self.row,column=16).value
        self.chart_months = self.sheet_cust.cell(row=self.row,column=15).value

        
        self.wb_cust.save('G:\\Finance software\\Database\\newLoan.xlsx')

        #Modifying Brokers data in the file named brokers basics
        self.wb_broker = load_workbook('G:\\Finance software\\Database\\BrokerBasics.xlsx')
        self.sheet_broker = self.wb_broker.active

        maxcol = self.sheet_broker.max_column
        maxrow = self.sheet_broker.max_row

        count = 0
        
        for i in range(1,maxrow+1):
            cell_obj = self.sheet_broker.cell(row=i,column=1)
            print(cell_obj.value)
            if(str(cell_obj.value)==self.broker_name):
                self.sheet_broker.cell(row=i,column=2).value -=int(self.old_amount)
                self.sheet_broker.cell(row=i,column=2).value +=int(self.updated_amount)
                self.sheet_broker.cell(row=i,column=6).value -=int(self.old_amount)
                self.sheet_broker.cell(row=i,column=6).value +=int(self.updated_amount)
                self.sheet_broker.cell(row=i,column=8).value -=int(self.old_amount)
                self.sheet_broker.cell(row=i,column=8).value +=int(self.updated_amount)
                count = 1
                self.wb_broker.save('G:\\Finance software\\Database\\BrokerBasics.xlsx') 
                break
        if(count == 0):
            print("No such Borker Found!!")


        #Modifying data in the particular persons file ie.chart of the particular person

        self.wb_chart = load_workbook('G:\\Finance software\\Database\\CustomerCharts\\'+str(self.unique_id)+'.xlsx')
        self.sheet_chart = self.wb_chart.active

        self.sheet_chart.cell(row=3,column=3).value = "Loan Amount :- "+self.updated_amount
        self.updated_month_emi = (float(self.updated_amount)*float(self.interest_rate))/int(self.chart_months)
        for i in range(9,9+int(self.chart_months)):
            self.sheet_chart.cell(row=i,column=2).value = self.updated_month_emi
        
        
        self.wb_chart.save("G:\\Finance software\\Database\\CustomerCharts\\"+str(self.unique_id)+".xlsx")
        
       

        
