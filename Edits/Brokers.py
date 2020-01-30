import tkinter as tk
import json
from openpyxl import *
import os
import time

class Edit(tk.Frame):
    def __init__(self, *args):
        tk.Frame.__init__(self, *args)
    def show(self):
        self.lift()
        
class CashBrok(Edit):
        
    def __init__(self, *args):
        
        Edit.__init__(self, *args)

        self.labeljName = tk.Label(self, text="Edit Brokers File",height=4)
        self.labeljName.grid(row=0,column=5)

        #--------------------------Veh No.-----------------------------------
        self.labelVehNo = tk.Label(self, text="Broker Name:-",height=4)
        self.labelVehNo.grid(row=1,column=1,padx=(200,0))
        self.e11 = tk.Entry(self,width=30)
        self.e11.grid(row=1,columnspan=2,column=2,padx=(50,0))
        self.tke11 = tk.StringVar()
        self.tke11.set("Nothing is done yet!")

        self.buttonFind = tk.Button(self, text="Search", fg="green",width=20)
        self.buttonFind.grid(row=1,column=4,padx=(100,0))
        self.buttonFind.config(command = self.searchBroker)
    

    def searchBroker(self):
        
         
        
        self.tke11.set(self.e11.get())
        self.setk11 = self.tke11.get()

        self.wb_brok = load_workbook('G:\\Finance software\\Database\\BrokerBasics.xlsx')

        self.sheet_brok = self.wb_brok.active
        
        self.maxcol = self.sheet_brok.max_column
        self.maxrow = self.sheet_brok.max_row
        c=0
        for i in range(1,self.maxrow+1):
            
            cell_obj = self.sheet_brok.cell(row=i,column=1)
            if(str(cell_obj.value)==str(self.setk11)):
                c=1
                
                self.row = i

                self.SetName = tk.Label(self, text="Modified Brokers Name:-",height=4)
                self.SetName.grid(row=2,column=1,padx=(200,0))
                self.b11 = tk.Entry(self,width=30)
                self.b11.grid(row=2,columnspan=2,column=2,padx=(50,0))

                self.SetAddress = tk.Label(self, text="Address:-",height=4)
                self.SetAddress.grid(row=3,column=1,padx=(200,0))
                self.b12 = tk.Entry(self,width=30)
                self.b12.grid(row=3,columnspan=2,column=2,padx=(50,0))

                self.SetCash = tk.Label(self, text="Amount in hand:-",height=4)
                self.SetCash.grid(row=4,column=1,padx=(200,0))
                self.b13 = tk.Entry(self,width=30)
                self.b13.grid(row=4,columnspan=2,column=2,padx=(50,0))
                

                self.tkb11 = tk.StringVar()  # create tk string variable
                self.tkb11.set('Nothing is done yet!')

                self.tkb12 = tk.StringVar()  # create tk string variable
                self.tkb12.set('Nothing is done yet!')

                self.tkb13 = tk.StringVar()  # create tk string variable
                self.tkb13.set('Nothing is done yet!')

                self.buttonFind = tk.Button(self, text="Submit", fg="green",width=20)
                self.buttonFind.grid(row=5,column=2,padx=(0,0))
                self.buttonFind.config(command = self.changeBrokerData)
                c=1
                break
            
        if(c==0):
            self.SetAmount = tk.Label(self, text="Sorry Couldnt Find the Broker! Check the broker name and enter again!!",height=4)
            self.SetAmount.grid(row=2,column=3,padx=(200,0))
        
    def changeBrokerData(self):
        self.tkb11.set(self.b11.get())
        self.tkb12.set(self.b12.get())
        self.tkb13.set(self.b13.get())
        
        self.setb11 = self.tkb11.get()
        self.setb12 = self.tkb12.get()
        self.setb13 = self.tkb13.get()

        self.sheet_brok.cell(row=self.row,column=1).value = str(self.setb11)
        self.sheet_brok.cell(row=self.row,column=9).value = str(self.setb12)
        self.sheet_brok.cell(row=self.row,column=2).value = str(self.setb13)

        self.wb_brok.save("G:\\Finance software\\Database\\BrokerBasics.xlsx")

            
