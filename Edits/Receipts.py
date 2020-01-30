import tkinter as tk
import json
from openpyxl import *
import os
import time

class Edit(tk.Frame):
    def __init__(self, *args):
        tk.Frame.__init__(self, *args)
    def show(self):
        self.lift()



class Receipt(Edit):
    def __init__(self, *args):
        
        Edit.__init__(self, *args)
                
        self.labeljName = tk.Label(self, text="Edit Invoices",height=4)
        self.labeljName.grid(row=0,column=3)

        #Take veh no, previous receipt amount, previous number of receipts paid, new data
        self.SetVehNo = tk.Label(self, text="VehNo:-",height=4)
        self.SetVehNo.grid(row=1,column=1,padx=(200,0))
        self.e11 = tk.Entry(self,width=30)
        self.e11.grid(row=1,columnspan=2,column=2)
        self.tke11 = tk.StringVar()
        self.tke11.set("Nothing is done yet!")

        self.buttonFind = tk.Button(self, text="Search", fg="green",width=20)
        self.buttonFind.grid(row=1,column=4,padx=(200,0))
        self.buttonFind.config(command = self.search_vehno)

        

    def search_vehno(self):
        self.tke11.set(self.e11.get())
        self.setk11 = self.tke11.get()


        self.wb_cust = load_workbook('G:\\Finance software\\Database\\newLoan.xlsx')

        self.sheet_cust = self.wb_cust.active

        maxcol = self.sheet_cust.max_column
        maxrow = self.sheet_cust.max_row
        c=0
        for i in range(1,maxrow+1):
            
            cell_obj = self.sheet_cust.cell(row = i,column = 14)
            print(str(cell_obj.value))
            if(str(cell_obj.value)==str(self.setk11)):
                print("dinve")

                
                self.row = i
                
                self.PrevInvoiceAmt = tk.Label(self, text="Previous Invoice Amount:-",height=4)
                self.PrevInvoiceAmt.grid(row=2,column=1,padx=(200,0))
                self.e12 = tk.Entry(self,width=30)
                self.e12.grid(row=2,columnspan=2,column=2)
                self.tke12 = tk.StringVar()
                self.tke12.set("Nothing is done yet!")
                
                self.NumPrevInvoice = tk.Label(self, text="Number of Month in previous Invoice:-",height=4)
                self.NumPrevInvoice.grid(row=2,column=4,padx=(200,0))
                self.e13 = tk.Entry(self,width=30)
                self.e13.grid(row=2,columnspan=2,column=5)
                self.tke13 = tk.StringVar()
                self.tke13.set("Nothing is done yet!")
                
                self.UpdatedAmount = tk.Label(self, text="Updated amount:-",height=4)
                self.UpdatedAmount.grid(row=3,column=1,padx=(200,0))
                self.e14 = tk.Entry(self,width=30)
                self.e14.grid(row=3,columnspan=2,column=2)
                self.tke14 = tk.StringVar()
                self.tke14.set("Nothing is done yet!")

                self.UpdatedMonths = tk.Label(self, text="Updated Number of Months:-",height=4)
                self.UpdatedMonths.grid(row=3,column=4,padx=(200,0))
                self.e15 = tk.Entry(self,width=30)
                self.e15.grid(row=3,columnspan=2,column=5)
                self.tke15 = tk.StringVar()
                self.tke15.set("Nothing is done yet!")

                self.buttonFind = tk.Button(self, text="Submit", fg="green",width=20)
                self.buttonFind.grid(row=4,column=4,padx=(0,0))
                self.buttonFind.config(command = self.changeInvoice)
                c=1
                break
            
        if(c==0):
            self.SetAmount = tk.Label(self, text="Sorry Couldnt Find the Vehicle! Check the Vehicle Number",height=4)
            self.SetAmount.grid(row=2,column=3,padx=(200,0))

    def changeInvoice(self):
        self.wb_cust = load_workbook('G:\\Finance software\\Database\\newLoan.xlsx')
        
        self.unique_id = self.sheet_cust.cell(row=self.row,column=21).value
        self.broker_name = self.sheet_cust.cell(row=self.row,column=11).value
        self.chart_months = self.sheet_cust.cell(row=self.row,column=15).value

        self.tke12.set(self.e12.get())
        self.setk12 = self.tke12.get()

        self.tke13.set(self.e13.get())
        self.setk13 = self.tke13.get()
        


        self.tke14.set(self.e14.get())
        self.setk14 = self.tke14.get()

        self.tke15.set(self.e15.get())
        self.setk15 = self.tke15.get()



        self.wb_broker = load_workbook('G:\\Finance software\\Database\\BrokerBasics.xlsx')
        self.sheet_broker = self.wb_broker.active

        maxcol = self.sheet_broker.max_column
        maxrow = self.sheet_broker.max_row

        
        count = 0

        
        for i in range(1,maxrow+1):
            cell_obj = self.sheet_broker.cell(row=i,column=1)
            print(cell_obj.value)
            if(str(cell_obj.value)==self.broker_name):
                self.sheet_broker.cell(row=i,column=2).value -=int(self.setk12)
                self.sheet_broker.cell(row=i,column=2).value +=int(self.setk15)
                count = 1
                self.wb_broker.save('G:\\Finance software\\Database\\BrokerBasics.xlsx') 
                break
        if(count == 0):
            print("No such Borker Found!!")


        self.wb_chart = load_workbook('G:\\Finance software\\Database\\CustomerCharts\\'+str(self.unique_id)+'.xlsx')
        self.sheet_chart = self.wb_chart.active
        c=0
        for i in range(9+int(self.chart_months)-1,8,-1):
            if(str(self.sheet_chart.cell(row=i,column=3).value)=="Paid" or str(self.sheet_chart.cell(row=i,column=3).value)=="Settled" and c==0):
                print(i)
                c+=1
            elif(str(self.sheet_chart.cell(row=i,column=3).value)=="Paid" or str(self.sheet_chart.cell(row=i,column=3).value)=="Settled" and c!=0 and c<int(self.setk13)):
                print(i)
                c+=1
            else:
                start = i
                break

        for i in range(start+1,start+int(self.setk15)+1):
            self.sheet_chart.cell(row=i,column=3).value = "Paid"
            print("Line",i)

        self.wb_chart.save("G:\\Finance software\\Database\\CustomerCharts\\"+str(self.unique_id)+".xlsx")







