import tkinter as tk
import json
import uuid
import os
import time
from openpyxl import *

class File(tk.Frame):
    def __init__(self, *args):
        tk.Frame.__init__(self, *args)
    def show(self):
        self.lift()

            
class NewInvoice(File):
    
    def __init__(self, *args):
        
       
        File.__init__(self, *args)
       

        self.labeljName = tk.Label(self, text="Generate Invoice and Add due to chart",height=4)
        self.labeljName.grid(row=0,column=3,padx=(200,0))

        self.SetVehNo = tk.Label(self, text="VehNo:-",height=4)
        self.SetVehNo.grid(row=1,column=1,padx=(200,0))
        self.e11 = tk.Entry(self,width=30)
        self.e11.grid(row=1,columnspan=2,column=2)
       
        self.tke11 = tk.StringVar()  # create tk string variable
        self.tke11.set('Nothing is done yet!')


        self.SetDate = tk.Label(self, text="Due Payment Date:-",height=4)
        self.SetDate.grid(row=2,column=1,padx=(200,0))
        self.e12 = tk.Entry(self,width=30)
        self.e12.grid(row=2,columnspan=2,column=2)
        
        self.tke12 = tk.StringVar()  # create tk string variable
        self.tke12.set('Nothing is done yet!')

        self.SetAmount = tk.Label(self, text="Due Amount:-",height=4)
        self.SetAmount.grid(row=3,column=1,padx=(200,0))
        self.e13 = tk.Entry(self,width=30)
        self.e13.grid(row=3,columnspan=2,column=2)
        
        self.tke13 = tk.StringVar()  # create tk string variable
        self.tke13.set('Nothing is done yet!')

        self.SetNo = tk.Label(self, text="No. of Dues:-",height=4)
        self.SetNo.grid(row=4,column=1,padx=(200,0))
        self.e14 = tk.Entry(self,width=30)
        self.e14.grid(row=4,columnspan=2,column=2)
        
        self.tke14 = tk.StringVar()  # create tk string variable
        self.tke14.set('Nothing is done yet!')

        self.buttonSettle = tk.Button(self, text="Generate Invoice", fg="green",width=20)
        self.buttonSettle.grid(row=5,column=2)
        self.buttonSettle.config(command = self.invoice)


    def invoice(self):

        self.tke11.set(self.e11.get())
        self.setk11 = self.tke11.get()

        
        self.tke12.set(self.e12.get())
        self.setk12 = self.tke12.get()

        self.tke13.set(self.e13.get())
        self.setk13 = self.tke13.get()

        self.tke14.set(self.e14.get())
        self.setk14 = self.tke14.get()



        self.wb_cust = load_workbook('G:\\Finance software\\Database\\newLoan.xlsx')

        self.sheet_custuid = self.wb_cust.active

            
        maxcol = self.sheet_custuid.max_column
        maxrow = self.sheet_custuid.max_row

        for i in range(1,maxrow+1):
            
            cell_obj = self.sheet_custuid.cell(row=i,column=14)
            print(str(self.setk11))
            if(str(cell_obj.value)==str(self.setk11)):
                temp = 0
                
                cell_obj_uid = self.sheet_custuid.cell(row=i,column=21)
                uid = str(cell_obj_uid.value)
                print("uid",uid)

                self.wb_file = load_workbook("G:\\Finance software\\Database\\CustomerCharts\\"+uid+".xlsx")
                self.sheet_file = self.wb_file.active

                maxrow = self.sheet_file.max_row

                cell_broker = str(self.sheet_file.cell(row=3,column=5).value)
                broker = cell_broker.lstrip('Broker :-')
                print(broker)

                for j in range(9,maxrow+1):
                    cell_obj_file = self.sheet_file.cell(row=j,column=3)
                    cell_obj_date = self.sheet_file.cell(row=j,column=4)
                    cell_obj_amount = self.sheet_file.cell(row=j,column=2)

                    print(cell_obj_file.value)

                    if(str(cell_obj_file.value) == "Unpaid" and temp<int(self.setk14)):

                        self.sheet_file.cell(row=j,column=3).value = "Paid"
                        self.sheet_file.cell(row=j,column=4).value = self.setk12
                        temp+=1

                self.wb_file.save("G:\\Finance software\\Database\\CustomerCharts\\"+str(uid)+".xlsx")

                    
                self.wb_broker = load_workbook('G:\\Finance software\\Database\\BrokerBasics.xlsx')
                self.sheet_f = self.wb_broker.active

                maxrow_broker = self.sheet_f.max_row

                for k in range(3,maxrow_broker+1):
                    cell_broker_name = self.sheet_f.cell(row=k,column=1).value
                    if(str(cell_broker_name)==broker):
                        self.sheet_f.cell(row=k,column=2).value+=int(self.setk13)
                        break
                self.wb_broker.save('G:\\Finance software\\Database\\BrokerBasics.xlsx')

                break
        self.cancel()

    def cancel(self):
        self.e11.delete(0,'end')
        self.e12.delete(0,'end')
        self.e13.delete(0,'end')
        self.e14.delete(0,'end')

        

        
        

       

       

       


       
