import tkinter as tk
import json
from Files.NewInvoices import NewInvoice
from Files.FileChart import CreateChart
from Files.SettleFile import Settle
from Files.FileBroker import NewBroker
import uuid
import os
from openpyxl import *
import time


class File(tk.Frame):
    def __init__(self, *args):
        tk.Frame.__init__(self, *args)
    def show(self):
        self.lift()


            
class NewLoan(File):
    def __init__(self, *args):
        File.__init__(self, *args)

        #--------------------------Excel self.sheet Creation------------------------------------
        self.wb = load_workbook('G:\\Finance software\\Database\\newLoan.xlsx')
        self.sheet = self.wb.active

        self.sheet.column_dimensions['A'].width = 30
        self.sheet.column_dimensions['B'].width = 30
        self.sheet.column_dimensions['C'].width = 20
        self.sheet.column_dimensions['D'].width = 50
        self.sheet.column_dimensions['E'].width = 30
        self.sheet.column_dimensions['F'].width = 30
        self.sheet.column_dimensions['G'].width = 20
        self.sheet.column_dimensions['H'].width = 50
        self.sheet.column_dimensions['I'].width = 20
        self.sheet.column_dimensions['J'].width = 20
        self.sheet.column_dimensions['K'].width = 30
        self.sheet.column_dimensions['L'].width = 15
        self.sheet.column_dimensions['M'].width = 15
        self.sheet.column_dimensions['N'].width = 20
        self.sheet.column_dimensions['O'].width = 20
        self.sheet.column_dimensions['P'].width = 20
        self.sheet.column_dimensions['Q'].width = 20
        self.sheet.column_dimensions['R'].width = 20
        self.sheet.column_dimensions['S'].width = 20
        

        # write given data to an excel spreadself.sheet 
        # at particular location 
        self.sheet.cell(row=1, column=1).value = "Customer Name"
        self.sheet.cell(row=1, column=2).value = "Father/Husbands Name"
        self.sheet.cell(row=1, column=3).value = "Phone Number"
        self.sheet.cell(row=1, column=4).value = "Address"
        self.sheet.cell(row=1, column=5).value = "Gaurentee Name"
        self.sheet.cell(row=1, column=6).value = "Gaurentee Fathers Name"
        self.sheet.cell(row=1, column=7).value = "Gaurentee Number"
        self.sheet.cell(row=1, column=8).value = "Gaurentee Address"
        self.sheet.cell(row=1, column=9).value = "Loan Date"
        self.sheet.cell(row=1, column=10).value = "Amount"
        self.sheet.cell(row=1, column=11).value = "Broker"
        self.sheet.cell(row=1, column=12).value = "Make"
        self.sheet.cell(row=1, column=13).value = "Model"
        self.sheet.cell(row=1, column=14).value = "Vehicle Number"
        self.sheet.cell(row=1, column=15).value = "Chart Months"
        self.sheet.cell(row=1, column=16).value = "Interest Rate"
        self.sheet.cell(row=1, column=17).value = "Deposit"
        self.sheet.cell(row=1, column=18).value = "Number of Documents"
        self.sheet.cell(row=1, column=19).value = "List of Documents"
        self.sheet.cell(row=1, column=20).value = "RC Number"
        self.sheet.cell(row=1, column=21).value = "Unique ID"
        
        



        self.labeljName = tk.Label(self, text="New Loan",height=4)
        self.labeljName.grid(row=0,column=3,padx=(200,0))

        #--------------------Line 1 - 3 inputs - name,fname,phno---------------------------
        self.labelName = tk.Label(self, text="Name:-",height=4)
        self.labelName.grid(row=1,column=1,padx=(100,0))
        self.e11 = tk.Entry(self,width=30)
        self.e11.grid(row=1,column=2)

        self.tke11 = tk.StringVar()  # create tk string variable
        self.tke11.set('Nothing is done yet!')

        
        self.labelFname = tk.Label(self, text="Fathers/Husbands Name:-")
        self.labelFname.grid(row=1,column=3,padx=(100,0))
        self.e12 = tk.Entry(self,width=30)
        self.e12.grid(row=1,column=4)

        self.tke12 = tk.StringVar()  # create tk string variable
        self.tke12.set('Nothing is done yet!')

        self.labelPhNo = tk.Label(self, text="Phone No:-")
        self.labelPhNo.grid(row=1,column=5,padx=(100,0))
        self.e13 = tk.Entry(self,width=30)
        self.e13.grid(row=1,column=6)

        self.tke13 = tk.StringVar()  # create tk string variable
        self.tke13.set('Nothing is done yet!')

        #------------------------Line 2 - 1 input - addr line 1,2---------------------------

        self.labelAddr1 = tk.Label(self, text="Address:-",height=4)
        self.labelAddr1.grid(row=4,column=1,padx=(200,0))
        self.e21 = tk.Entry(self,width=60)
        self.e21.grid(row=4,column=2,columnspan=2)

        self.tke21 = tk.StringVar()  # create tk string variable
        self.tke21.set('Nothing is done yet!')

        #----------------------------Line 3 - Guarantee Details----------------------------

        self.labelGName = tk.Label(self, text="Gaureenter Name:-",height=4)
        self.labelGName.grid(row=3,column=1,padx=(100,0))
        self.e31 = tk.Entry(self,width=30)
        self.e31.grid(row=3,column=2)

        self.tke31 = tk.StringVar()  # create tk string variable
        self.tke31.set('Nothing is done yet!')
        
        self.labelGFname = tk.Label(self, text="Gaureenter Fathers/Husbands Name:-")
        self.labelGFname.grid(row=3,column=3,padx=(100,0))
        self.e32 = tk.Entry(self,width=30)
        self.e32.grid(row=3,column=4)

        self.tke32 = tk.StringVar()  # create tk string variable
        self.tke32.set('Nothing is done yet!')

        self.labelGPhNo = tk.Label(self, text="Gaureenter Phone No:-")
        self.labelGPhNo.grid(row=3,column=5,padx=(100,0))
        self.e33 = tk.Entry(self,width=30)
        self.e33.grid(row=3,column=6)

        self.tke33 = tk.StringVar()  # create tk string variable
        self.tke33.set('Nothing is done yet!')

        #------------------------Line 4 - 2 input - Gaddr line 1,2---------------------------

        self.labelGAddr1 = tk.Label(self, text="Gaureenter Address:-",height=4)
        self.labelGAddr1.grid(row=4,column=4)
        self.e41 = tk.Entry(self,width=60)
        self.e41.grid(row=4,column=5,columnspan=2)

        self.tke41 = tk.StringVar()  # create tk string variable
        self.tke41.set('Nothing is done yet!')

        #----------------------------Line 5 - Loan Details----------------------------

        self.labelDate = tk.Label(self, text="Loan Date:-",height=4)
        self.labelDate.grid(row=5,column=1,padx=(100,0))
        self.e51 = tk.Entry(self,width=30)
        self.e51.grid(row=5,column=2)

        self.tke51 = tk.StringVar()  # create tk string variable
        self.tke51.set('Nothing is done yet!')
        
        self.labelAmount = tk.Label(self, text="Amount:-")
        self.labelAmount.grid(row=5,column=3,padx=(100,0))
        self.e52 = tk.Entry(self,width=30)
        self.e52.grid(row=5,column=4)

        self.tke52 = tk.StringVar()  # create tk string variable
        self.tke52.set('Nothing is done yet!')

        self.labelBroker = tk.Label(self, text="Broker:-")
        self.labelBroker.grid(row=5,column=5,padx=(100,0))
        self.e53 = tk.Entry(self,width=30)
        self.e53.grid(row=5,column=6)

        self.tke53 = tk.StringVar()  # create tk string variable
        self.tke53.set('Nothing is done yet!')

        #----------------------------Line 6 - Vehicle Info----------------------------

        self.labelMake = tk.Label(self, text="Make:-",height=4)
        self.labelMake.grid(row=6,column=1,padx=(100,0))
        self.e61 = tk.Entry(self,width=30)
        self.e61.grid(row=6,column=2)

        self.tke61 = tk.StringVar()  # create tk string variable
        self.tke61.set('Nothing is done yet!')
        
        self.labelModel = tk.Label(self, text="Model:-")
        self.labelModel.grid(row=6,column=3,padx=(100,0))
        self.e62 = tk.Entry(self,width=30)
        self.e62.grid(row=6,column=4)

        self.tke62 = tk.StringVar()  # create tk string variable
        self.tke62.set('Nothing is done yet!')

        self.labelVehNo = tk.Label(self, text="Vehicle No:-")
        self.labelVehNo.grid(row=6,column=5,padx=(100,0))
        self.e63 = tk.Entry(self,width=30)
        self.e63.grid(row=6,column=6)

        self.tke63 = tk.StringVar()  # create tk string variable
        self.tke63.set('Nothing is done yet!')

        #----------------------------Line 7 - Inputs on Loan----------------------------

        self.labelChart = tk.Label(self, text="Chart Months:-",height=4)
        self.labelChart.grid(row=7,column=1,padx=(100,0))
        self.e71 = tk.Entry(self,width=30)
        self.e71.grid(row=7,column=2)

        self.tke71 = tk.StringVar()  # create tk string variable
        self.tke71.set('Nothing is done yet!')
        
        self.labelInterest = tk.Label(self, text="Interest Rate:-")
        self.labelInterest.grid(row=7,column=3,padx=(100,0))
        self.e72 = tk.Entry(self,width=30)
        self.e72.grid(row=7,column=4)

        self.tke72 = tk.StringVar()  # create tk string variable
        self.tke72.set('Nothing is done yet!')

        self.labelDepo = tk.Label(self, text="Deposit:-")
        self.labelDepo.grid(row=7,column=5,padx=(100,0))
        self.e73 = tk.Entry(self,width=30)
        self.e73.grid(row=7,column=6)

        self.tke73 = tk.StringVar()  # create tk string variable
        self.tke73.set('Nothing is done yet!')


        #----------------------------Line 8 - Documents----------------------------

        self.labelNoDocs = tk.Label(self, text="No.of Docs:-",height=4)
        self.labelNoDocs.grid(row=8,column=1,padx=(100,0))
        self.e81 = tk.Entry(self,width=30)
        self.e81.grid(row=8,column=2)

        self.tke81 = tk.StringVar()  # create tk string variable
        self.tke81.set('Nothing is done yet!')

        self.labelDocs = tk.Label(self, text="Documents List:-",height=4)
        self.labelDocs.grid(row=9,column=1,padx=(200,0))
        self.e82 = tk.Entry(self,width=60)
        self.e82.grid(row=9,column=2,columnspan=2)

        self.tke82 = tk.StringVar()  # create tk string variable
        self.tke82.set('Nothing is done yet!')

        #----------------------------Line 9 - RC.No----------------------------

        self.labelRC = tk.Label(self, text="RC No:-",height=4)
        self.labelRC.grid(row=10,column=1,padx=(100,0))
        self.e91 = tk.Entry(self,width=30)
        self.e91.grid(row=10,column=2)

        self.tke91 = tk.StringVar()  # create tk string variable
        self.tke91.set('Nothing is done yet!')

        

        #------------------------------Buttons---------------------------------
        self.buttonSave = tk.Button(self, text="SAVE", fg="green",width=20)
        self.buttonSave.grid(row=11,column=3)
        self.buttonSave.config(command = self.savedata)

        self.buttonCancel = tk.Button(self, text="CANCEL", fg="red",width=20)
        self.buttonCancel.grid(row=11,column=4)
        self.buttonCancel.config(command = self.canceldata)

        #------------------------Opening File in append mode--------------------

        #self.FileOpen = open("Database/LoanFile/loanfile.txt","a+") 
        
        self.d={}


    
        

    def savedata(self):

        #--------------------------------------------------Getting all the entries and saving in a particular variable and thereby saving in database---------------------------------------
        
        self.tke11.set(self.e11.get())     
        self.k11=self.tke11.get()
        self.d["custname"]=self.k11

        self.tke12.set(self.e12.get())
        self.k12 = self.tke12.get()
        self.d["custfname"]=self.k12

        
        self.tke13.set(self.e13.get())
        self.k13 = self.tke13.get()
        self.d["custphno"]=self.k13
        
        self.tke21.set(self.e21.get())
        self.k21 = self.tke21.get()
        self.d["add1"]=self.k21
        
        
        
        self.tke31.set(self.e31.get())
        self.k31 = self.tke31.get()
        self.d["gauname"]=self.k31
        
        self.tke32.set(self.e32.get())
        self.k32 = self.tke32.get()
        self.d["gaufname"]=self.k32
        
        self.tke33.set(self.e33.get())
        self.k33 = self.tke33.get()
        self.d["gauphno"]=self.k33
        
        self.tke41.set(self.e41.get())
        self.k41 = self.tke41.get()
        self.d["gadd1"]=self.k41
        
        
        
        self.tke53.set(self.e53.get())
        self.k53 = self.tke53.get()
        self.d["broker"]=self.k53
        
        self.tke52.set(self.e52.get())
        self.k52 = self.tke52.get()
        self.d["amount"]=self.k52
        
        self.tke51.set(self.e51.get())
        self.k51 = self.tke51.get()
        self.d["ldate"]=self.k51

        
        self.tke61.set(self.e61.get())
        self.k61 = self.tke61.get()
        self.d["make"]=self.k61
        
        self.tke62.set(self.e62.get())
        self.k62 = self.tke62.get()
        self.d["model"]=self.k62
        
        self.tke63.set(self.e63.get())
        self.k63 = self.tke63.get()
        self.d["vehno"]=self.k63
        
        self.tke71.set(self.e71.get())
        self.k71 = self.tke71.get()
        self.d["cmonths"]=self.k71

        self.tke72.set(self.e72.get())
        self.k72 = self.tke72.get()
        self.d["irate"]=self.k72
        
        self.tke73.set(self.e73.get())
        self.k73 = self.tke73.get()
        self.d["deposit"]=self.k73
        
        self.tke81.set(self.e81.get())
        self.k81 = self.tke81.get()
        self.d["nodocs"]=self.k81

        self.tke82.set(self.e82.get())
        self.k82 = self.tke82.get()
        self.d["docslist"]=self.k82
        
        self.tke91.set(self.e91.get())
        self.k91 = self.tke91.get()
        self.d["rcno"]=self.k91

        #--------------------------------------------------saving data in excel sheet names newLoan----------------------------------------

        current_row = self.sheet.max_row 
        current_column = self.sheet.max_column



        

        self.sheet.cell(row=current_row + 1, column=1).value = self.k11
        self.sheet.cell(row=current_row + 1, column=2).value = self.k12
        self.sheet.cell(row=current_row + 1, column=3).value = self.k13
        self.sheet.cell(row=current_row + 1, column=4).value = self.k21
        self.sheet.cell(row=current_row + 1, column=5).value = self.k31
        self.sheet.cell(row=current_row + 1, column=6).value = self.k32
        self.sheet.cell(row=current_row + 1, column=7).value = self.k33
        self.sheet.cell(row=current_row + 1, column=8).value = self.k41
        self.sheet.cell(row=current_row + 1, column=9).value = self.k51
        self.sheet.cell(row=current_row + 1, column=10).value = self.k52
        self.sheet.cell(row=current_row + 1, column=11).value = self.k53

        self.sheet.cell(row=current_row + 1, column=12).value = self.k61
        self.sheet.cell(row=current_row + 1, column=13).value = self.k62
        self.sheet.cell(row=current_row + 1, column=14).value = self.k63
        self.sheet.cell(row=current_row + 1, column=15).value = self.k71
        self.sheet.cell(row=current_row + 1, column=16).value = self.k72
        self.sheet.cell(row=current_row + 1, column=17).value = self.k73
        self.sheet.cell(row=current_row + 1, column=18).value = self.k81
        self.sheet.cell(row=current_row + 1, column=19).value = self.k82
        self.sheet.cell(row=current_row + 1, column=20).value = self.k91

        self.wb.save('G:\\Finance software\\Database\\newLoan.xlsx') 

        

        self.fileOpn()

        
        
        #-------------------------------------Writing json file by converting it from dictionary to a file--------------------------------------------
        self.FileOpen.write(json.dumps(self.d))
        self.FileOpen.write("\n")
        self.FileOpen.close()
        self.canceldata()
        self.e11.delete(0, 'end')
        self.popupmsg("Data Saved !!")

        #---------------------------Data successfully saved---------------------------------


    def fileOpn(self):

        self.months = ["Jan","Feb","Mar","Apr","May","Jun","July","Aug","Sep","Oct","Nov","Dec"]
        self.d2={}

        #------------------------------Dates segeregated as per date,month,year for easy opn-----------------------------------------
        self.datek51 = self.k51.split(".")
        self.k51month = self.datek51[1]
        self.k51date = self.datek51[0]
        self.k51year = self.datek51[2]

        #----------------------------------Empty list declared------------------------------------

        self.l=[]

        #--------------------creating chart months-------------------------
        self.startmonth = int(self.k51month)%12+1
        self.l.append(self.startmonth)
        self.tempmonth = int(self.startmonth)
        self.totalmonths = int(self.k71)
        self.i=self.totalmonths
        while(self.i>1):
            self.l.append(self.tempmonth%12+1)
            self.tempmonth+=1
            self.i-=1

        #----------------------------------------------------------------------------------

        self.l2=[]
        self.upcomingdueyear=int(self.k51year)
        self.upcomingduemonth = int(self.k51month)+1
        if(self.upcomingduemonth>12):
            self.upcomingduemonth%=12
            self.upcomingdueyear+=1

        self.upcomingduedate = int(self.k51date)
        self.l2.append(self.upcomingduedate)
        self.l2.append(self.upcomingduemonth)
        self.l2.append(self.upcomingdueyear)

        
        self.l3=[]
        i = 0
        self.upcomingdueyear=int(self.k51year)
        self.upcomingduemonth = int(self.k51month)+1
        while(len(self.l3)!=len(self.l)):
            if(self.upcomingduemonth>12):
                self.upcomingduemonth%=12
                self.upcomingdueyear+=1

            self.upcomingduedate = int(self.k51date)
            self.l3.append([])
            self.l3[i].append(self.upcomingduedate)
            self.l3[i].append(self.upcomingduemonth)
            self.l3[i].append(self.upcomingdueyear)
            self.upcomingduemonth+=1
            i+=1
            

        
        
        print(self.l2)
        print(self.l)
        print(self.l3)
        self.d2["Name"]=self.k11
        self.d2["vehno"]=self.k63
        self.d2["chartMonths"]=self.l
        self.d2["upcomingdue"]=self.l2
        self.d2["ChartStatus"]="Ongoing"

        #-----------------------------Creating unique id for all customers--------------------
        self.unique=""
        self.uid = uuid.uuid1()
        self.unique+=str(self.uid)
        print(self.unique)

        self.CustFile = open("Database/Customers/"+self.unique+".txt","a+")
        self.CustToUid = open("Database/CustToUid.txt","a+")

        
        current_row = self.sheet.max_row 
        current_column = self.sheet.max_column
        self.sheet.cell(row=current_row, column=21).value = self.unique

        self.wb.save('G:\\Finance software\\Database\\newLoan.xlsx')
        



        #--------------------------Creating charts-------------------------------------
        self.CustFile.write(json.dumps(self.d2))
        self.CustFile.write("/n")
        self.CustFile.close()

        #---------------------------Maintaining customer record with their unique chart numbers---------------------------
        self.d3={}
        self.d3["uid"]=self.unique
        self.d3["vehno"]=self.k63

        self.CustToUid.write(json.dumps(self.d3))
        self.CustToUid.write("/n")

        self.CustToUid.close()

        self.newChart()
        self.brokerData(self.k53,self.k52)
        
        
        


    def canceldata(self):

        
        print(self.k11)
        #----------------------Clearing all the entries---------------------------
        self.e11.delete(0, 'end')
        self.e12.delete(0, 'end')
        self.e13.delete(0, 'end')
        self.e21.delete(0, 'end')
        
        self.e31.delete(0, 'end')
        self.e32.delete(0, 'end')
        self.e33.delete(0, 'end')
        self.e41.delete(0, 'end')
        
        self.e51.delete(0, 'end')
        self.e52.delete(0, 'end')
        self.e53.delete(0, 'end')
        self.e61.delete(0, 'end')
        self.e62.delete(0, 'end')
        self.e63.delete(0, 'end')
        self.e71.delete(0, 'end')
        self.e72.delete(0, 'end')
        self.e73.delete(0, 'end')
        self.e81.delete(0, 'end')
        self.e82.delete(0, 'end')
        self.e91.delete(0, 'end')

        #----------------------------since the file was closed, reopening the file------------------------------------
        
        print("cancelled")
        self.FileOpen = open("Database/LoanFile/loanfile.txt","a")


        
    def popupmsg(self,msg):
        LARGE_FONT= ("Verdana", 12)
        NORM_FONT = ("Helvetica", 10)
        SMALL_FONT = ("Helvetica", 8)
        popup = tk.Tk()
        popup.wm_title("!")
        label = tk.Label(popup, text=msg, font=NORM_FONT)
        label.pack(side="top", fill="x", pady=10)
        B1 = tk.Button(popup, text="Okay", command = popup.destroy)
        B1.pack()
        popup.mainloop()




    def brokerFileSetup(self):
        self.wb = load_workbook('G:\\Finance software\\Database\\BrokerBasics.xlsx')
        self.sheet_broker = self.wb.active

        self.sheet_broker.column_dimensions['A'].width = 30
        self.sheet_broker.column_dimensions['B'].width = 30
        self.sheet_broker.column_dimensions['C'].width = 30
        self.sheet_broker.column_dimensions['D'].width = 30
        self.sheet_broker.column_dimensions['E'].width = 30
        self.sheet_broker.column_dimensions['F'].width = 30
        self.sheet_broker.column_dimensions['G'].width = 30
        self.sheet_broker.column_dimensions['H'].width = 30
        self.sheet_broker.column_dimensions['I'].width = 30
        
        

        # write given data to an excel spreadself.sheet 
        # at particular location 
        self.sheet_broker.cell(row=1, column=1).value = "Broker Name"
        self.sheet_broker.cell(row=1, column=2).value = "Amount In Hand"
        self.sheet_broker.cell(row=1, column=3).value = "No. of Customers"
        self.sheet_broker.cell(row=1, column=4).value = "Settled Customers"
        self.sheet_broker.cell(row=1, column=5).value = "Total Customers"
        self.sheet_broker.cell(row=1, column=6).value = "Loan Given"
        self.sheet_broker.cell(row=1, column=7).value = "Loan Settled"
        self.sheet_broker.cell(row=1, column=8).value = "Loan till date"
        self.sheet_broker.cell(row=1, column=9).value = "Broker Address"



        
    def brokerData(self,brokername,loan_amount):
        self.wb_broker = load_workbook('G:\\Finance software\\Database\\BrokerBasics.xlsx')
        self.sheet_broker = self.wb_broker.active

        maxcol = self.sheet_broker.max_column
        maxrow = self.sheet_broker.max_row

        count = 0
        
        for i in range(1,maxrow+1):
            cell_obj = self.sheet_broker.cell(row=i,column=1)
            print(cell_obj.value)
            if(str(cell_obj.value)==brokername):
                self.sheet_broker.cell(row=i, column=2).value = int(self.sheet_broker.cell(row=i, column=2).value) - int(loan_amount)
                self.sheet_broker.cell(row=i, column=3).value +=1
                self.sheet_broker.cell(row=i, column=5).value +=1
                self.sheet_broker.cell(row=i, column=6).value +=int(loan_amount)
                self.sheet_broker.cell(row=i, column=8).value +=int(loan_amount)
                count = 1
                self.wb_broker.save('G:\\Finance software\\Database\\BrokerBasics.xlsx') 
                break
        if(count == 0):
            print("No such Borker Found!!")
        
                
                


        
    
     

    def newChart(self):
        book = Workbook()
        sheet = book.active
        sheet['A1'] = "Name :- "+self.k11
        sheet['C1'] = "Ph.No :- "+self.k13
        sheet['A2'] = "Address :- "+self.k21
        sheet['A3'] = "Loan Date :- "+self.k51
        sheet['C3'] = "Loan Amount :- "+self.k52
        sheet['E3'] = "Broker :- "+self.k53
        sheet['A4'] = "Veh. No :- "+self.k63
        sheet['A5'] = "Chart Months :- "+self.k71
        sheet['C5'] = "Interest Rate :- "+self.k72
        sheet['E5'] = "Deposit :- "+self.k73
        sheet['A6'] = "RC - No :- "+self.k91

        sheet['A8'] = "Months"
        sheet['B8'] = "Amount Per Month"
        sheet['C8'] = "Status"
        sheet['D8'] = "Date"


        uid = self.unique
        l = self.l3
        a = "A"
        b = "B"
        c = "C"
        d = "D"
        j = 0
        for i in range(9,len(l)+9):
            sheet[a+str(i)] = str(l[j][0])+"."+str(l[j][1])+"."+str(l[j][2])
            sheet[b+str(i)] = (int(self.k52)*float(self.k72))/int(self.k71)
            sheet[c+str(i)] = "Unpaid"
            sheet[d+str(i)] = "Unpaid"
            j+=1
            

        
        d = 1
        book.save("G:\\Finance software\\Database\\CustomerCharts\\"+str(uid)+".xlsx")   



    def monthDataFileSetup(self,monthYear):
        book = workbook()
        sheet = book.active

        
        d = 1

        
        


        
    def monthDataEdit(self,month,year):
        self.wbMonthDataCheck = load_workbook('G:\\Finance software\\Database\\AvailableMonthData.xlsx')
        self.CheckSheet = self.wbMonthDataCheck.active

        maxrow = CheckSheet.max_row
        count = 0

        monthYear = month+str(year)

        for i in range(1,maxrow+1):
            cell_obj_month = self.CheckSheet.cell(row=i,col=1)
            cell_obj_year = self.CheckSheet.cell(row=i,col=2)

            if(cell_obj_year==year and cell_obj_month == month):
                count = 1
                break
            
        if(count==1):
            self.wbMonthData = load_workbook('G:\\Finance software\\Database\\MonthReports\\',monthYear,'.xlsx')


        else:
            self.CheckSheet.cell(row=maxrow+1, column=1).value = month
            self.CheckSheet.cell(row=maxrow+1,column=2).value = year


            #----------------creating new file by name of monthYear--------------------------
            monthDataFileSetup(monthYear)

            
        








