import tkinter as tk
import json
import uuid
import os
from openpyxl import *


class File(tk.Frame):
    def __init__(self, *args):
        tk.Frame.__init__(self, *args,bg="black")
    def show(self):
        self.lift()
            
class Settle(File):
    def __init__(self, *args):
        File.__init__(self, *args)


        self.labeljName = tk.Label(self, text="New Loan",bg="black",fg="white",height=4)
        self.labeljName.grid(row=0,column=3,padx=(200,0))

        self.SetVehNo = tk.Label(self, text="VehNo:-",bg="black",fg="white",height=4)
        self.SetVehNo.grid(row=1,column=1,padx=(200,0))
        self.e11 = tk.Entry(self,width=30)
        self.e11.grid(row=1,columnspan=2,column=2)
        
        self.tke11 = tk.StringVar()  # create tk string variable
        self.tke11.set('Nothing is done yet!')


        self.SetDate = tk.Label(self, text="Settling Date:-",bg="black",fg="white",height=4)
        self.SetDate.grid(row=2,column=1,padx=(200,0))
        self.e12 = tk.Entry(self,width=30)
        self.e12.grid(row=2,columnspan=2,column=2)
        
        self.tke12 = tk.StringVar()  # create tk string variable
        self.tke12.set('Nothing is done yet!')

        self.SetAmount = tk.Label(self, text="Settling Amount:-",bg="black",fg="white",height=4)
        self.SetAmount.grid(row=3,column=1,padx=(200,0))
        self.e13 = tk.Entry(self,width=30)
        self.e13.grid(row=3,columnspan=2,column=2)
        
        self.tke13 = tk.StringVar()  # create tk string variable
        self.tke13.set('Nothing is done yet!')



        self.buttonSettle = tk.Button(self, text="SETTLE", fg="green",width=20)
        self.buttonSettle.grid(row=1,column=4)
        self.buttonSettle.config(command = self.settle)


        

    def settle(self):

        self.tke11.set(self.e11.get())
        self.setk11 = self.tke11.get()

        
        self.tke12.set(self.e12.get())
        self.setk12 = self.tke12.get()

        self.tke13.set(self.e13.get())
        self.setk13 = self.tke13.get()

        self.wb_cust = load_workbook('G:\\Finance software\\Database\\newLoan.xlsx')

        self.sheet_custuid = self.wb_cust.active

            
        maxcol = self.sheet_custuid.max_column
        maxrow = self.sheet_custuid.max_row

        for i in range(1,maxrow+1):
            
            cell_obj = self.sheet_custuid.cell(row=i,column=14)
            print(str(self.setk11))
            if(str(cell_obj.value)==str(self.setk11)):
                
                cell_obj_uid = self.sheet_custuid.cell(row=i,column=21)
                uid = str(cell_obj_uid.value)
                print("uid",uid)

                self.wb_file = load_workbook("G:\\Finance software\\Database\\CustomerCharts\\"+uid+".xlsx")
                self.sheet_file = self.wb_file.active

                maxrow = self.sheet_file.max_row

                cell_broker = str(self.sheet_file.cell(row=3,column=5).value)
                broker = cell_broker.lstrip('Broker :-')
                print(broker)

                for j in range(9,maxrow+1):
                    cell_obj_file = self.sheet_file.cell(row=j,column=3)
                    cell_obj_date = self.sheet_file.cell(row=j,column=4)
                    cell_obj_amount = self.sheet_file.cell(row=j,column=2)

                    print(cell_obj_file.value)

                    if(str(cell_obj_file.value) == "Unpaid"):

                        self.sheet_file.cell(row=j,column=3).value = "Settled"
                        self.sheet_file.cell(row=j,column=4).value = self.setk12

                self.wb_file.save("G:\\Finance software\\Database\\CustomerCharts\\"+str(uid)+".xlsx")

                    
                self.wb_broker = load_workbook('G:\\Finance software\\Database\\BrokerBasics.xlsx')
                self.sheet_f = self.wb_broker.active

                maxrow_broker = self.sheet_f.max_row

                for k in range(3,maxrow_broker+1):
                    cell_broker_name = self.sheet_f.cell(row=k,column=1).value
                    if(str(cell_broker_name)==broker):
                        self.sheet_f.cell(row=k,column=2).value+=int(self.setk13)
                        self.sheet_f.cell(row=k,column=4).value+=1
                        self.sheet_f.cell(row=k,column=7).value+=int(self.setk13)
                        break
                self.wb_broker.save('G:\\Finance software\\Database\\BrokerBasics.xlsx')

                break

        

    def settlefile(self):
        

        
        print(self.setk11)
        print(self.setk11)
        self.FileOpen = open("Database/CustToUid.txt","r")
        self.f1=self.FileOpen.readlines()
        for i in self.f1:
            i=eval(i)
            print(i["vehno"])
            if(i["vehno"]==str(self.setk11)):
                print("Yeah")
                self.custid = i["uid"]
                self.FileOpenCust = open("Database/Customers/"+self.custid+".txt","r+")
                self.readlinesf1 = self.FileOpenCust.readlines()
                self.d={}
                for j in self.readlinesf1:
                    print(j)
                    j=eval(j)
                    self.d["Name"] = j["Name"]
                    self.d["vehno"] = j["vehno"]
                    self.d["chartMonths"] = j["chartMonths"]
                    self.d["upcomingdue"] = j["upcomingdue"]
                    self.d["ChartStatus"] = "Settled"
                self.FileOpenCust.close()
                self.FileOpenCust = open("Database/Customers/"+self.custid+".txt","w")

                self.FileOpenCust.write(json.dumps(self.d))
                self.FileOpenCust.write("/n")
                self.FileOpenCust.close()


        
                
                    
                    


                        

                    
                            
                            
                    

            
            
        
            
            
            

        

        










