import tkinter as tk
from openpyxl import *
import time
import os

class File(tk.Frame):
    def __init__(self, *args):
        tk.Frame.__init__(self, *args)
    def show(self):
        self.lift()


class NewBroker(File):
    
    
    def __init__(self, *args):
        
        File.__init__(self, *args)
       
        self.labeljName = tk.Label(self, text="New Broker",height=4)
        self.labeljName.grid(row=0,column=4)

        self.SetName = tk.Label(self, text="Name:-",height=4)
        self.SetName.grid(row=1,column=3,padx=(200,0))
        self.b11 = tk.Entry(self,width=30)
        self.b11.grid(row=1,columnspan=2,column=4,padx=(50,0))

        self.SetAddress = tk.Label(self, text="Address:-",height=4)
        self.SetAddress.grid(row=2,column=3,padx=(200,0))
        self.b12 = tk.Entry(self,width=30)
        self.b12.grid(row=2,columnspan=2,column=4,padx=(50,0))

        self.SetCash = tk.Label(self, text="Cash Given:-",height=4)
        self.SetCash.grid(row=3,column=3,padx=(200,0))
        self.b13 = tk.Entry(self,width=30)
        self.b13.grid(row=3,columnspan=2,column=4,padx=(50,0))

       
        
        self.tkb11 = tk.StringVar()  # create tk string variable
        self.tkb11.set('Nothing is done yet!')

        self.tkb12 = tk.StringVar()  # create tk string variable
        self.tkb12.set('Nothing is done yet!')

        self.tkb13 = tk.StringVar()  # create tk string variable
        self.tkb13.set('Nothing is done yet!')

        

        self.buttonSettle = tk.Button(self, text="ADD", fg="green",width=20)
        self.buttonSettle.grid(row=4,column=4,padx=(0,0))
        self.buttonSettle.config(command = self.newBroker)

    def newBroker(self):
        self.tkb11.set(self.b11.get())
        self.tkb12.set(self.b12.get())
        self.tkb13.set(self.b13.get())

        self.setb11 = self.tkb11.get()
        self.setb12 = self.tkb12.get()
        self.setb13 = self.tkb13.get()

        self.wb_broker = load_workbook('G:\\Finance software\\Database\\BrokerBasics.xlsx')
        self.sheet_broker = self.wb_broker.active

        self.sheet_broker.column_dimensions['A'].width = 30
        self.sheet_broker.column_dimensions['B'].width = 30
        self.sheet_broker.column_dimensions['C'].width = 30
        self.sheet_broker.column_dimensions['D'].width = 30
        self.sheet_broker.column_dimensions['E'].width = 30
        self.sheet_broker.column_dimensions['F'].width = 30
        self.sheet_broker.column_dimensions['G'].width = 30
        self.sheet_broker.column_dimensions['H'].width = 30
        self.sheet_broker.column_dimensions['I'].width = 30
        
        

        # write given data to an excel spreadself.sheet 
        # at particular location 
        self.sheet_broker.cell(row=1, column=1).value = "Broker Name"
        self.sheet_broker.cell(row=1, column=2).value = "Amount In Hand"
        self.sheet_broker.cell(row=1, column=3).value = "No. of Customers"
        self.sheet_broker.cell(row=1, column=4).value = "Settled Customers"
        self.sheet_broker.cell(row=1, column=5).value = "Total Customers"
        self.sheet_broker.cell(row=1, column=6).value = "Loan Given"
        self.sheet_broker.cell(row=1, column=7).value = "Loan Settled"
        self.sheet_broker.cell(row=1, column=8).value = "Loan till date"
        self.sheet_broker.cell(row=1, column=9).value = "Broker Address"

        maxcol = self.sheet_broker.max_column
        i = self.sheet_broker.max_row

        self.sheet_broker.cell(row=i+1, column=1).value = self.setb11
        self.sheet_broker.cell(row=i+1, column=2).value = self.setb13
        self.sheet_broker.cell(row=i+1, column=3).value = 0
        self.sheet_broker.cell(row=i+1, column=4).value = 0
        self.sheet_broker.cell(row=i+1, column=5).value = 0
        self.sheet_broker.cell(row=i+1, column=6).value = 0
        self.sheet_broker.cell(row=i+1, column=7).value = 0
        self.sheet_broker.cell(row=i+1, column=8).value = 0
        self.sheet_broker.cell(row=i+1, column=9).value = self.setb12

        self.wb_broker.save('G:\\Finance software\\Database\\BrokerBasics.xlsx')


         
